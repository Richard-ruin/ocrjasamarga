# app/services/excel_service.py - Updated with enhanced features + Time in H22:J22
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from pathlib import Path
from typing import List
from datetime import datetime
from PIL import Image
import os
import logging
import tempfile
import uuid

logger = logging.getLogger(__name__)

def format_indonesian_date(date_str: str) -> str:
    """
    Format tanggal ke format Indonesia: "6 November 2023"
    """
    try:
        # Parse tanggal dari format ISO (YYYY-MM-DD)
        if isinstance(date_str, str) and len(date_str) >= 10:
            date_obj = datetime.strptime(date_str[:10], "%Y-%m-%d")
        else:
            date_obj = datetime.now()
        
        # Mapping bulan Indonesia
        bulan_indonesia = {
            1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
            5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
            9: "September", 10: "Oktober", 11: "November", 12: "Desember"
        }
        
        # Format: "6 November 2023"
        formatted_date = f"{date_obj.day} {bulan_indonesia[date_obj.month]} {date_obj.year}"
        return formatted_date
        
    except Exception as e:
        logger.warning(f"Error formatting date {date_str}: {e}")
        # Fallback ke tanggal hari ini
        today = datetime.now()
        bulan_indonesia = {
            1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
            5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
            9: "September", 10: "Oktober", 11: "November", 12: "Desember"
        }
        return f"{today.day} {bulan_indonesia[today.month]} {today.year}"

def format_time_only(time_str: str) -> str:
    """
    Format waktu ke format HH:MM tanpa simbol ':'
    """
    try:
        # Jika time_str berupa datetime string lengkap
        if isinstance(time_str, str) and len(time_str) >= 8:
            # Coba parse dari format datetime lengkap
            if 'T' in time_str:
                time_part = time_str.split('T')[1][:8]  # Ambil bagian waktu
                time_obj = datetime.strptime(time_part, "%H:%M:%S").time()
            elif ' ' in time_str and ':' in time_str:
                # Format "YYYY-MM-DD HH:MM:SS"
                time_part = time_str.split(' ')[1][:8]
                time_obj = datetime.strptime(time_part, "%H:%M:%S").time()
            else:
                # Langsung format HH:MM:SS atau HH:MM
                time_obj = datetime.strptime(time_str[:5], "%H:%M").time()
        else:
            # Fallback ke waktu sekarang
            time_obj = datetime.now().time()
        
        # Format: "08.30" (tanpa simbol :)
        formatted_time = f"{time_obj.hour:02d}.{time_obj.minute:02d}"
        return formatted_time
        
    except Exception as e:
        logger.warning(f"Error formatting time {time_str}: {e}")
        # Fallback ke waktu sekarang
        now = datetime.now()
        return f"{now.hour:02d}.{now.minute:02d}"

def generate_excel(data: List[dict], save_dir: Path) -> Path:
    """
    Generate Excel file dengan koordinat dari cache data dan enhanced features
    """
    template_path = Path("uploads/template.xlsx")
    
    if not template_path.exists():
        raise FileNotFoundError(f"Template Excel tidak ditemukan: {template_path}")
    
    wb = load_workbook(template_path)
    ws = wb.active
    start_row = 9

    logger.info(f"=== EXCEL SERVICE START (Enhanced with Time) ===")
    logger.info(f"Generating Excel with {len(data)} entries")

    # ✅ ENHANCED FEATURE 1: Format tanggal jadwal untuk J4
    try:
        # Ambil tanggal dari jadwal (tanggal_inspeksi dari data)
        tanggal_jadwal = ""
        if data and len(data) > 0:
            # Prioritas: tanggal_inspeksi > tanggal > created_at
            tanggal_jadwal = data[0].get("tanggal_inspeksi", "")
            if not tanggal_jadwal:
                tanggal_jadwal = data[0].get("tanggal", "")
            if not tanggal_jadwal:
                tanggal_jadwal = data[0].get("created_at", "")
        
        if not tanggal_jadwal:
            tanggal_jadwal = datetime.now().isoformat()
        
        formatted_date = format_indonesian_date(tanggal_jadwal)
        ws['J4'] = f": {formatted_date}"
        logger.info(f"✅ SET J4 (Tanggal Jadwal) = ': {formatted_date}' from '{tanggal_jadwal}'")
        
        
    except Exception as date_error:
        logger.error(f"Error setting jadwal date in J4: {date_error}")
        # Fallback
        today_formatted = format_indonesian_date(datetime.now().isoformat())
        ws['J4'] = f": {today_formatted}"

    # ✅ ENHANCED FEATURE 2: Nama Aset di merged cells C2:H3
    try:
        # Ambil nama aset dari data entry pertama
        nama_aset = ""
        if data and len(data) > 0:
            nama_aset = data[0].get("nama_aset", "")
            if not nama_aset:
                # Fallback ke ID aset jika nama tidak ada
                nama_aset = data[0].get("id_aset", "")
        
        if not nama_aset:
            nama_aset = "Aset Tidak Diketahui"
        
        # Set di cell C2 (yang di-merge dengan C2:H3)
        ws['C2'] = nama_aset
        logger.info(f"✅ SET C2 (Merged C2:H3) = '{nama_aset}'")
        
    except Exception as aset_error:
        logger.error(f"Error setting nama aset in C2: {aset_error}")
        # Fallback
        ws['C2'] = "Nama Aset"

    # ✅ ENHANCED FEATURE 4: Waktu di merged cells H22:J22 (BARU!)
    try:
        # Ambil waktu dari jadwal
        waktu_jadwal = ""
        if data and len(data) > 0:
            # Prioritas: waktu_inspeksi > waktu > created_at
            waktu_jadwal = data[0].get("waktu_inspeksi", "")
            if not waktu_jadwal:
                waktu_jadwal = data[0].get("waktu", "")
            if not waktu_jadwal:
                waktu_jadwal = data[0].get("created_at", "")
        
        if not waktu_jadwal:
            waktu_jadwal = datetime.now().isoformat()
        
        formatted_time = format_time_only(waktu_jadwal)
        ws['H22'] = formatted_time  # Set di H22 (yang di-merge dengan H22:J22)
        logger.info(f"✅ SET H22 (Merged H22:J22) = '{formatted_time}' from '{waktu_jadwal}'")
        
    except Exception as time_error:
        logger.error(f"Error setting waktu in H22: {time_error}")
        # Fallback ke waktu sekarang
        now_formatted = format_time_only(datetime.now().isoformat())
        ws['H22'] = now_formatted

    # ✅ BUAT TEMP DIRECTORY UNTUK GAMBAR
    temp_dir = Path(tempfile.gettempdir()) / "excel_images"
    temp_dir.mkdir(exist_ok=True)
    
    temp_files_to_cleanup = []  # Track files to cleanup

    try:
        for i, entry in enumerate(data):
            row = start_row + i
            
            logger.info(f"=== PROCESSING ROW {row} (Entry {i+1}) ===")
            
            # Basic data
            no_value = entry.get("no", i + 1)
            jalur_value = entry.get("jalur", "")
            
            ws[f'B{row}'] = no_value
            ws[f'C{row}'] = jalur_value

            # ✅ KOORDINAT - Langsung dari cache
            latitude = entry.get("latitude")
            longitude = entry.get("longitude")
            
            logger.info(f"  Coordinates from cache: lat='{latitude}', lon='{longitude}'")
            
            # Set latitude
            if latitude and str(latitude).strip():
                lat_clean = str(latitude).strip()
                ws[f'D{row}'] = lat_clean
                logger.info(f"  ✅ SET D{row} = '{lat_clean}'")
            else:
                ws[f'D{row}'] = ""
                logger.warning(f"  ❌ D{row} empty")
                
            # Set longitude
            if longitude and str(longitude).strip():
                lon_clean = str(longitude).strip()
                ws[f'E{row}'] = lon_clean
                logger.info(f"  ✅ SET E{row} = '{lon_clean}'")
            else:
                ws[f'E{row}'] = ""
                logger.warning(f"  ❌ E{row} empty")

            # Kondisi checkboxes
            kondisi = entry.get("kondisi", "").lower()
            ws[f'F{row}'] = "√" if kondisi == "baik" else ""
            ws[f'G{row}'] = "√" if kondisi == "sedang" else ""
            ws[f'H{row}'] = "√" if kondisi == "buruk" else ""

            # Keterangan
            keterangan_value = entry.get("keterangan", "")
            ws[f'I{row}'] = keterangan_value

            # ✅ FOTO - Perbaiki path temporary
            foto_path = entry.get("foto_path")
            if foto_path and os.path.exists(str(foto_path)):
                try:
                    logger.info(f"  Processing image: {foto_path}")
                    
                    # Buka dan resize gambar
                    img = Image.open(str(foto_path))
                    img.thumbnail((400, 300), Image.Resampling.LANCZOS)
                    
                    # ✅ BUAT PATH TEMPORARY YANG BENAR
                    temp_filename = f"temp_img_{uuid.uuid4().hex}.png"
                    img_temp_path = temp_dir / temp_filename
                    
                    logger.info(f"  Saving temp image to: {img_temp_path}")
                    
                    # Save temporary image
                    img.save(str(img_temp_path), "PNG")
                    temp_files_to_cleanup.append(img_temp_path)
                    
                    # Add to Excel
                    excel_img = ExcelImage(str(img_temp_path))
                    excel_img.width = 120
                    excel_img.height = 90
                    ws.add_image(excel_img, f'J{row}')
                    
                    logger.info(f"  ✅ Image added to J{row}")
                    
                except Exception as e:
                    logger.error(f"  ❌ Image error for row {row}: {e}")
            else:
                logger.warning(f"  No valid image for row {row}: foto_path='{foto_path}'")

        # ✅ ENHANCED FEATURE 3: Log informasi lengkap yang ditambahkan
        logger.info("=== ENHANCED FEATURES APPLIED ===")
        logger.info(f"1. Tanggal Jadwal in J4: {ws['J4'].value}")
        logger.info(f"2. Asset name in C2: {ws['C2'].value}")
        logger.info(f"3. Waktu in H22: {ws['H22'].value}")  # NEW!
        logger.info(f"4. Total data rows processed: {len(data)}")
        
        # Informasi tambahan dari data jadwal
        if data and len(data) > 0:
            sample_entry = data[0]
            logger.info("=== JADWAL & ASET INFO ===")
            logger.info(f"Jadwal ID: {sample_entry.get('jadwal_id', 'N/A')}")
            logger.info(f"Tanggal Jadwal: {sample_entry.get('tanggal_inspeksi', 'N/A')}")
            logger.info(f"Waktu Jadwal: {sample_entry.get('waktu_inspeksi', 'N/A')}")
            logger.info(f"Nama Inspektur: {sample_entry.get('nama_inspektur', 'N/A')}")
            logger.info(f"Alamat Inspeksi: {sample_entry.get('alamat_inspeksi', 'N/A')}")
            logger.info(f"ID Aset: {sample_entry.get('id_aset', 'N/A')}")
            logger.info(f"Nama Aset: {sample_entry.get('nama_aset', 'N/A')}")
            logger.info(f"Jenis Aset: {sample_entry.get('jenis_aset', 'N/A')}")
            logger.info(f"Lokasi Aset: {sample_entry.get('lokasi_aset', 'N/A')}")

        # Save file
        save_dir.mkdir(parents=True, exist_ok=True)
        filename = f"output-enhanced-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        save_path = save_dir / filename
        
        logger.info(f"Saving Enhanced Excel to: {save_path}")
        wb.save(save_path)
        
        logger.info(f"✅ Enhanced Excel generation completed: {save_path}")
        return save_path
        
    finally:
        # ✅ CLEANUP TEMPORARY FILES
        for temp_file in temp_files_to_cleanup:
            try:
                if temp_file.exists():
                    temp_file.unlink()
                    logger.info(f"Cleaned up temp file: {temp_file}")
            except Exception as cleanup_error:
                logger.warning(f"Failed to cleanup {temp_file}: {cleanup_error}")